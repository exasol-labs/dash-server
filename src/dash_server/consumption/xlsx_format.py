"""Streaming, typed, spreadsheet-safe XLSX formatter with a provenance sheet."""

from __future__ import annotations

from datetime import date, datetime, time
from decimal import Decimal
import hashlib
import json
from pathlib import Path
import re
from collections.abc import Callable, Iterable
from typing import Any

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter

from dash_server.exceptions import DashServerError

from .csv_format import limit_error


_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\']")
_MIN_COLUMN_WIDTH = 12
_MAX_COLUMN_WIDTH = 40


def write_xlsx(
    path: Path,
    *,
    columns: list[str],
    batches: Iterable[list[list[Any]]],
    max_rows: int,
    max_bytes: int,
    cancelled: Callable[[], bool],
    provenance: dict[str, Any],
    on_progress: Callable[[int, int], None] | None = None,
) -> dict[str, Any]:
    workbook = Workbook(write_only=True)
    data_sheet = workbook.create_sheet(_safe_sheet_name(provenance.get("output_id")))
    data_sheet.freeze_panes = "A2"
    data_sheet.auto_filter.ref = f"A1:{get_column_letter(max(1, len(columns)))}1"
    for index, column in enumerate(columns, start=1):
        data_sheet.column_dimensions[get_column_letter(index)].width = _bounded_width(column)
    data_sheet.append([_safe_cell(data_sheet, name) for name in columns])
    row_count = 0
    try:
        for batch in batches:
            if cancelled():
                raise DashServerError(
                    category="consumption_job_cancelled",
                    summary="Export cancellation was requested.",
                    details={},
                    jsonrpc_code=-32032,
                    http_status=409,
                )
            for row in batch:
                if row_count >= max_rows:
                    raise limit_error("rows", max_rows)
                data_sheet.append([_safe_cell(data_sheet, value) for value in row])
                row_count += 1
            if on_progress is not None:
                on_progress(row_count, 0)
        provenance_sheet = workbook.create_sheet("Provenance")
        provenance_sheet.column_dimensions["A"].width = 24
        provenance_sheet.column_dimensions["B"].width = 60
        for key, value in {**provenance, "row_count": row_count, "limit_outcome": "within_limits"}.items():
            rendered = json.dumps(value, sort_keys=True) if isinstance(value, (dict, list)) else value
            provenance_sheet.append([_safe_cell(provenance_sheet, key), _safe_cell(provenance_sheet, rendered)])
        workbook.save(path)
    finally:
        workbook.close()
    byte_size = path.stat().st_size
    if byte_size > max_bytes:
        raise limit_error("bytes", max_bytes)
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return {"row_count": row_count, "byte_size": byte_size, "sha256": digest.hexdigest()}


def _safe_cell(sheet: Any, value: Any) -> Any:
    """Return a typed cell; strings are pinned to the string type so no source
    value can ever be interpreted as an Excel formula."""
    if value is None:
        return None
    if isinstance(value, bool | int | float | Decimal | datetime | date | time):
        cell = WriteOnlyCell(sheet, value=value)
        return cell
    cell = WriteOnlyCell(sheet, value=str(value))
    cell.data_type = "s"
    return cell


def _safe_sheet_name(candidate: Any) -> str:
    name = _INVALID_SHEET_CHARS.sub(" ", str(candidate or "")).strip()
    return name[:31] or "Data"


def _bounded_width(header: str) -> int:
    return min(_MAX_COLUMN_WIDTH, max(_MIN_COLUMN_WIDTH, len(header) + 2))


__all__ = ["write_xlsx"]
