from __future__ import annotations

import json
from dataclasses import replace
from pathlib import Path
import re
import sqlite3
import sys
from threading import Event
import time
from types import SimpleNamespace
from typing import Any

import pytest

from dash_server.auth import AuthContext, Principal
from dash_server.app_factory import create_app
from dash_server.consumption import consumption_contract_hash, normalize_consumption_contract
from dash_server.consumption.execution import DatasetStream, ExasolDatasetExecutor
from dash_server.exceptions import DashServerError

from _helpers import wait_for


_APP_PY = """from dash import Dash, html


def create_dash_app(server, url_base_pathname, metadata):
    prefix = url_base_pathname.rstrip(\"/\") + \"/\"
    app = Dash(
        __name__,
        server=server,
        routes_pathname_prefix=\"/\",
        requests_pathname_prefix=prefix,
        title=metadata.get(\"title\", \"Consumption Test\"),
    )
    app.layout = html.Div([html.H1(metadata.get(\"title\", \"Consumption Test\"))])
    return app
"""


def _call_tool(
    client,
    name: str,
    arguments: dict[str, Any],
    request_id: int = 1,
    *,
    headers: dict[str, str] | None = None,
):
    return client.post(
        "/mcp",
        headers=headers,
        json={
            "jsonrpc": "2.0",
            "id": request_id,
            "method": "tools/call",
            "params": {"name": name, "arguments": arguments},
        },
    )


def _read_resource(client, uri: str, request_id: int = 2) -> dict[str, Any]:
    response = client.post(
        "/mcp",
        json={
            "jsonrpc": "2.0",
            "id": request_id,
            "method": "resources/read",
            "params": {"uri": uri},
        },
    )
    assert response.status_code == 200
    return json.loads(response.get_json()["result"]["contents"][0]["text"])


def _consumption_contract(*, source_path: str = "queries/export.sql") -> dict[str, Any]:
    return {
        "outputs": [
            {
                "id": "monthly-close-detail",
                "title": "Monthly close detail",
                "description": "Posting detail for one accounting period.",
                "kind": "dataset",
                "source": {
                    "type": "exasol_sql",
                    "data_source": "primary",
                    "path": source_path,
                },
                "parameters": {
                    "type": "object",
                    "properties": {
                        "period": {
                            "type": "string",
                            "pattern": "^[0-9]{4}-[0-9]{2}$",
                        }
                    },
                    "required": ["period"],
                    "additionalProperties": False,
                },
                "formats": ["csv", "xlsx"],
                "classification": "confidential",
                "limits": {"max_rows": 5000, "max_bytes": 1000000},
                "allow_subscriptions": True,
                "allow_alerts": False,
            }
        ]
    }


class _ConsumptionSmokeFakeConnection:
    """A pyexasol connection stand-in that accepts any query and returns no rows.

    PS26-BUG-005: `create_app`/`app_build` now always preflight the first revision,
    which runs the `sql_smoke` probe against the bound Exasol profile. None of these
    tests' queries reference a bad column, so a bare "the query parsed" success is
    all the fixture needs - it does not need to fake real result rows.
    """

    def execute(self, sql_text: str, params: dict[str, object] | None = None) -> object:
        return None

    def close(self) -> None:
        return None


class _ConsumptionSmokeFakePyExasolModule:
    def connect(self, **kwargs: object) -> _ConsumptionSmokeFakeConnection:
        return _ConsumptionSmokeFakeConnection()


def _wire_fake_exasol_connector(app) -> None:
    app.extensions["exasol_dashboard_service"].connection_manager.connector_loader = (
        lambda: _ConsumptionSmokeFakePyExasolModule()
    )


def _create_analytics_prod_profile(client, *, request_id: int = 500) -> None:
    response = _call_tool(
        client,
        "exasol_profile_create_local",
        {
            "name": "analytics-prod",
            "backend": "onprem",
            "credential_mode": "password",
            "dsn": "demodb.exasol.com:8563",
            "user": "sys",
            "secret_value": "super-secret",
        },
        request_id=request_id,
    )
    assert response.status_code == 200


def _create_output_app(app, client, *, name: str = "finance-outputs", include_sql: bool = True):
    _wire_fake_exasol_connector(app)
    _create_analytics_prod_profile(client)
    files = [{"path": "app.py", "content": _APP_PY}]
    if include_sql:
        files.append(
            {
                "path": "queries/export.sql",
                "content": "SELECT {period!s} AS PERIOD FROM DUAL\n",
            }
        )
        files.append(
            {
                "path": "queries/sql_smoke.json",
                "content": json.dumps({"queries/export.sql": {"period": "2026-07"}}) + "\n",
            }
        )
    response = _call_tool(
        client,
        "app_create_from_files",
        {
            "name": name,
            "title": "Finance Outputs",
            "data_sources": {"primary": {"kind": "exasol", "profile": "analytics-prod"}},
            "consumption": _consumption_contract(),
            "files": files,
        },
    )
    assert response.status_code == 200
    result = response.get_json()["result"]
    assert result["isError"] is False
    return result["structuredContent"]


class _FakeDatasetExecutor:
    def __init__(self, rows: list[list[Any]] | None = None) -> None:
        self.rows = rows or [["2026-07", "=SUM(A1:A2)"], ["2026-07", "safe"]]
        self.preflight_calls = 0

    def preflight(self, revision, output) -> None:
        self.preflight_calls += 1

    def stream(self, revision, output, parameters, *, cancelled) -> DatasetStream:
        return DatasetStream(columns=["PERIOD", "VALUE"], batches=iter([[*self.rows]]))


class _BlockingDatasetExecutor(_FakeDatasetExecutor):
    def __init__(self) -> None:
        super().__init__()
        self.started = Event()
        self.release = Event()

    def stream(self, revision, output, parameters, *, cancelled) -> DatasetStream:
        def batches():
            self.started.set()
            self.release.wait(timeout=5)
            yield self.rows

        return DatasetStream(columns=["PERIOD", "VALUE"], batches=batches())


def _enable_phase1(app, executor=None):
    service = app.extensions["consumption_service"]
    service.policy = replace(service.policy, exports_enabled=True)
    service.policy_version = service.policy.version
    service.executor = executor or _FakeDatasetExecutor()
    return service


def _wait_for_job(client, job_id: str, expected: set[str] | None = None):
    expected = expected or {"succeeded", "failed", "cancelled"}
    deadline = time.monotonic() + 5
    while time.monotonic() < deadline:
        response = _call_tool(client, "export_get", {"job_id": job_id})
        payload = response.get_json()["result"]["structuredContent"]
        if payload["job"]["status"] in expected:
            return payload
        time.sleep(0.02)
    raise AssertionError(f"Export {job_id} did not reach {expected}")


@pytest.mark.slow
def test_phase0_output_discovery_matches_mcp_resource_and_ui(app, client):
    _create_output_app(app, client)

    tool_response = _call_tool(
        client,
        "app_outputs_list",
        {"name": "finance-outputs"},
        request_id=3,
    )
    tool_payload = tool_response.get_json()["result"]["structuredContent"]
    resource_payload = _read_resource(
        client,
        "dash://apps/finance-outputs/outputs",
        request_id=4,
    )
    ui_response = client.get("/manage/apps/finance-outputs/consumption")
    catalog_response = client.get("/")
    hosted_layout = client.get("/apps/finance-outputs/_dash-layout")

    tool_domain_payload = {key: value for key, value in tool_payload.items() if key != "guidance"}
    assert tool_domain_payload == resource_payload
    assert tool_payload["output_count"] == 1
    assert tool_payload["outputs"][0]["id"] == "monthly-close-detail"
    # Assert the discovery-phase behaviour this test is about rather than the
    # whole policy dict (exact limits/blocked-formats are pinned by the
    # dedicated phase/policy tests); the cross-surface parity check above
    # already guarantees the MCP and resource payloads carry an identical
    # policy shape.
    policy = tool_payload["outputs"][0]["policy"]
    assert policy["enabled"] is True
    assert policy["phase"] == "discovery_only"
    assert policy["executable"] is False
    assert policy["reason"] == "no_executable_format"
    assert policy["effective_formats"] == ["csv", "xlsx"]
    assert policy["format_availability"]["csv"]["reason"] == "exports_disabled"
    assert policy["format_availability"]["xlsx"]["reason"] == "exports_disabled"
    revision = app.extensions["registry"].get_current_revision("finance-outputs")
    assert revision is not None
    declared_consumption = revision.manifest["consumption"]
    assert declared_consumption["outputs"][0]["id"] == "monthly-close-detail"
    assert "policy" not in declared_consumption["outputs"][0]
    assert consumption_contract_hash(declared_consumption) == tool_payload["contract_hash"]
    assert revision.manifest["consumption_contract_hash"] == tool_payload["contract_hash"]
    assert ui_response.status_code == 200
    assert b"Monthly close detail" in ui_response.data
    assert b"Output discovery is available now" in ui_response.data
    assert b"Outputs (1)" in catalog_response.data
    assert b"/manage/apps/finance-outputs/consumption" in catalog_response.data
    assert hosted_layout.status_code == 200
    hosted_layout_text = json.dumps(hosted_layout.get_json())
    assert "/manage/apps/finance-outputs/consumption" in hosted_layout_text
    assert "__dash-server-exports-link" in hosted_layout_text


@pytest.mark.slow
def test_output_get_and_execution_context_are_revision_and_principal_bound(app, client):
    _create_output_app(app, client)
    response = _call_tool(
        client,
        "app_output_get",
        {"name": "finance-outputs", "output_id": "monthly-close-detail"},
    )
    payload = response.get_json()["result"]["structuredContent"]
    context = app.extensions["consumption_service"].execution_context(
        "finance-outputs",
        "monthly-close-detail",
        AuthContext.for_mode("local", auth_enabled=False),
    )

    assert payload["output"]["id"] == "monthly-close-detail"
    assert context.app_name == "finance-outputs"
    assert context.revision_number == payload["revision"]["revision_number"]
    assert context.output_contract_hash == payload["contract_hash"]
    assert context.principal_id == "local-admin"
    assert context.policy_version.startswith("consumption-v1:")


@pytest.mark.slow
def test_workspace_validation_fails_when_declared_output_source_is_missing(client):
    response = _call_tool(
        client,
        "app_create_from_files",
        {
            "name": "missing-output-source",
            "title": "Missing Output Source",
            "data_sources": {"primary": {"kind": "exasol", "profile": "analytics-prod"}},
            "consumption": _consumption_contract(),
            "files": [{"path": "app.py", "content": _APP_PY}],
        },
    )
    result = response.get_json()["result"]
    validation = result["structuredContent"]["error"]["details"]["validation"]

    assert result["isError"] is True
    assert result["structuredContent"]["error"]["category"] == "workspace_validation_error"
    assert validation["is_valid"] is False
    assert validation["consumption"]["status"] == "failed"
    assert validation["consumption"]["issues"][0]["path"] == "queries/export.sql"


def test_ps26_bug014_app_validate_warns_early_when_exports_are_disabled_server_wide(app, client) -> None:
    """PS26-BUG-014 regression: previously an agent could author a valid
    consumption.outputs contract, pass app_validate cleanly, build, and deploy live -
    only discovering exports are disabled server-wide (DASH_SERVER_CONSUMPTION_EXPORTS_ENABLED,
    default False) at the very last step, when app_export_create fails. app_validate now
    surfaces this as a non-blocking warning as early as validate/create time.
    """

    _create_output_app(app, client, name="finance-outputs-early-warning")

    response = _call_tool(client, "app_validate", {"name": "finance-outputs-early-warning"})
    result = response.get_json()["result"]
    assert result["isError"] is False
    report = result["structuredContent"]["validation"]

    assert report["consumption"]["status"] == "passed_with_warnings"
    assert report["is_valid"] is True
    warning = report["consumption"]["issues"][0]
    assert warning["level"] == "warning"
    assert "exports are disabled server-wide" in warning["message"]
    assert "consumption_exports_disabled" in warning["message"]

    summary = result["structuredContent"]["validation_summary"]
    assert summary["warning_count"] >= 1


def test_ps27_bug009_app_export_create_accepts_view_formats_at_the_schema_level(app, client) -> None:
    """PS27-BUG-009 regression: `app_export_create`'s format schema used to hardcode
    `enum: ["csv", "xlsx"]`, so a "view" kind output's own advertised
    `allowed_formats` (pdf/png/pptx, per `app_outputs_list`'s policy) were structurally
    unreachable through the only tool that could ever act on them - rejected with a
    generic `tool_validation_error` before the already-correct, format-availability-
    aware check in `ConsumptionService.create_export` ever ran. The schema must now
    accept every format the consumption contract supports, and delegate "is this
    format actually executable for this output" to that existing, more specific check.
    """

    create_response = _call_tool(
        client,
        "app_create_from_files",
        {
            "name": "view-output-app",
            "title": "View Output App",
            "consumption": {
                "outputs": [
                    {
                        "id": "dashboard-snapshot",
                        "title": "Dashboard snapshot",
                        "description": "A PDF snapshot of the live dashboard.",
                        "kind": "view",
                        "source": {"type": "dash_route", "path": "/"},
                        "parameters": {
                            "type": "object",
                            "properties": {},
                            "required": [],
                            "additionalProperties": False,
                        },
                        "formats": ["pdf", "png"],
                        "classification": "internal",
                    }
                ]
            },
            "files": [{"path": "app.py", "content": _APP_PY}],
        },
    )
    assert create_response.get_json()["result"]["isError"] is False

    outputs = _call_tool(client, "app_outputs_list", {"name": "view-output-app"}).get_json()["result"][
        "structuredContent"
    ]
    format_availability = outputs["outputs"][0]["policy"]["format_availability"]
    assert format_availability["pdf"]["reason"] in {"exports_disabled", "renderer_not_available"}

    export_response = _call_tool(
        client,
        "app_export_create",
        {"name": "view-output-app", "output_id": "dashboard-snapshot", "format": "pdf", "parameters": {}},
    )
    result = export_response.get_json()["result"]
    assert result["isError"] is True
    error = result["structuredContent"]["error"]
    # The request must fail for a *governance* reason (exports off, or the view
    # renderer not yet implemented) - never a schema-level "'pdf' is not one of
    # ['csv', 'xlsx']" rejection, which is what this bug produced unconditionally.
    assert error["category"] in {"consumption_exports_disabled", "consumption_format_unavailable"}
    assert error["category"] != "tool_validation_error"


def test_contract_rejects_unsafe_source_and_unknown_parameter_schema_features():
    data_sources = {"primary": {"kind": "exasol", "profile": "analytics-prod"}}
    with pytest.raises(DashServerError) as unsafe_path:
        normalize_consumption_contract(
            _consumption_contract(source_path="../secret.sql"),
            data_sources=data_sources,
        )
    assert unsafe_path.value.category == "consumption_contract_validation_error"
    assert unsafe_path.value.details["field"].endswith("source.path")

    contract = _consumption_contract()
    contract["outputs"][0]["parameters"]["properties"]["period"]["oneOf"] = []
    with pytest.raises(DashServerError) as unknown_schema:
        normalize_consumption_contract(contract, data_sources=data_sources)
    assert unknown_schema.value.category == "consumption_contract_validation_error"
    assert "Unknown field" in unknown_schema.value.summary


@pytest.mark.slow
def test_output_discovery_requires_app_export_authorization(app, client):
    _create_output_app(app, client)
    service = app.extensions["consumption_service"]
    registry = app.extensions["registry"]
    principal = Principal.authenticated_user(
        issuer="test",
        subject="viewer-1",
        email="viewer@example.test",
        roles=(),
        email_verified=True,
    )
    hosted_context = AuthContext(
        mode="hosted",
        auth_enabled=True,
        provider="trusted_proxy",
        principal=principal,
    )

    with pytest.raises(DashServerError) as denied:
        service.list_outputs("finance-outputs", hosted_context)
    assert denied.value.category == "consumption_authorization_denied"

    registry.grant_app_access(
        "finance-outputs",
        principal_type="user",
        principal_id=principal.principal_id,
        role="viewer",
        scope="all",
        created_by_principal_id="test-owner",
    )
    allowed = service.list_outputs("finance-outputs", hosted_context)
    assert allowed["authorization"]["allowed"] is True
    assert allowed["authorization"]["effective_role"] == "viewer"


@pytest.mark.slow
def test_consumption_registry_tables_and_phase1_columns_are_initialized(app):
    db_path = app.config["REGISTRY_DB_PATH"]
    with sqlite3.connect(db_path) as connection:
        tables = {
            row[0] for row in connection.execute("SELECT name FROM sqlite_master WHERE type = 'table'").fetchall()
        }
        job_columns = {row[1] for row in connection.execute("PRAGMA table_info(consumption_jobs)").fetchall()}

    assert {
        "consumption_jobs",
        "consumption_artifacts",
        "consumption_subscriptions",
        "consumption_alerts",
        "consumption_delivery_attempts",
        "consumption_audit_events",
    } <= tables
    assert {
        "context_json",
        "output_json",
        "parameters_redacted_json",
        "effective_limits_json",
        "cancel_requested_at",
        "expires_at",
    } <= job_columns


@pytest.mark.slow
class TestHostedViewerOutputCatalogParity:
    """Decomposition of ``test_hosted_viewer_has_same_authorized_output_
    catalog_through_mcp_and_ui``.

    The create -> grant -> list (MCP + UI) -> export -> download -> revoke
    end-to-end path runs once in the class-scoped ``flow`` fixture (it polls a
    background export job, hence stays ``slow``); each focused test asserts one
    authorization surface against the captured payloads.
    """

    @staticmethod
    @pytest.fixture(scope="class")
    def flow(tmp_path_factory):
        tmp_path = tmp_path_factory.mktemp("hosted_outputs")
        hosted_app = create_app(
            {
                "TESTING": True,
                "REGISTRY_DB_PATH": str(tmp_path / "registry.sqlite3"),
                "ARTIFACTS_ROOT": str(tmp_path / "artifacts"),
                "WORKSPACES_ROOT": str(tmp_path / "workspaces"),
                "DIAGNOSTICS_ROOT": str(tmp_path / "diagnostics"),
                "DEPENDENCY_STATE_ROOT": str(tmp_path / "dependency_state"),
                "GITOPS_REPO_PATH": str(tmp_path / "gitops-repo"),
                "EXASOL_SECRETS_ROOT": str(tmp_path / "exasol-secrets"),
                "AUTO_INSTALL_DEPENDENCIES": False,
                "PYTHON_EXECUTABLE": sys.executable,
                "DASH_SERVER_MODE": "hosted",
                "SECRET_KEY": "test-secret-key",
                "SESSION_COOKIE_SECURE": True,
                "SESSION_COOKIE_HTTPONLY": True,
                "SESSION_COOKIE_SAMESITE": "Lax",
                "DASH_SERVER_PUBLIC_BASE_URL": "https://dash.example.test",
                "DASH_SERVER_AUTH_PROVIDER": "trusted_proxy",
                "DASH_SERVER_TRUSTED_PROXY_HEADERS_ENABLED": True,
                "DASH_SERVER_TRUSTED_PROXY_ALLOWED_CIDRS": ("127.0.0.1/32",),
                "DASH_SERVER_BOOTSTRAP_ADMIN_PRINCIPAL_IDS": ("trusted_proxy:admin-1",),
                "DASH_SERVER_ALLOW_UNSAFE_INPROCESS": True,
            }
        )
        hosted_client = hosted_app.test_client()
        admin_headers = {
            "X-Forwarded-User": "admin-1",
            "X-Forwarded-Email": "admin@example.test",
        }
        viewer_headers = {
            "X-Forwarded-User": "viewer-1",
            "X-Forwarded-Email": "viewer@example.test",
        }
        stranger_headers = {
            "X-Forwarded-User": "stranger-1",
            "X-Forwarded-Email": "stranger@example.test",
        }
        _wire_fake_exasol_connector(hosted_app)
        profile_created = _call_tool(
            hosted_client,
            "exasol_profile_create_local",
            {
                "name": "analytics-prod",
                "backend": "onprem",
                "credential_mode": "password",
                "dsn": "demodb.exasol.com:8563",
                "user": "sys",
                "secret_value": "super-secret",
            },
            headers=admin_headers,
        )
        assert profile_created.status_code == 200
        files = [
            {"path": "app.py", "content": _APP_PY},
            {"path": "queries/export.sql", "content": "SELECT {period!s} AS PERIOD FROM DUAL\n"},
            {
                "path": "queries/sql_smoke.json",
                "content": json.dumps({"queries/export.sql": {"period": "2026-07"}}) + "\n",
            },
        ]
        created = _call_tool(
            hosted_client,
            "app_create_from_files",
            {
                "name": "hosted-outputs",
                "data_sources": {"primary": {"kind": "exasol", "profile": "analytics-prod"}},
                "consumption": _consumption_contract(),
                "files": files,
            },
            headers=admin_headers,
        )
        hosted_app.extensions["registry"].grant_app_access(
            "hosted-outputs",
            principal_type="user",
            principal_id="trusted_proxy:viewer-1",
            role="viewer",
            scope="all",
            created_by_principal_id="trusted_proxy:admin-1",
        )

        mcp_response = _call_tool(
            hosted_client,
            "app_outputs_list",
            {"name": "hosted-outputs"},
            headers=viewer_headers,
        )
        ui_response = hosted_client.get(
            "/manage/apps/hosted-outputs/consumption",
            headers=viewer_headers,
        )
        denied_mcp = _call_tool(
            hosted_client,
            "app_outputs_list",
            {"name": "hosted-outputs"},
            headers=stranger_headers,
        )
        denied_ui = hosted_client.get(
            "/manage/apps/hosted-outputs/consumption",
            headers=stranger_headers,
        )

        _enable_phase1(hosted_app)
        export_created = _call_tool(
            hosted_client,
            "app_export_create",
            {
                "name": "hosted-outputs",
                "output_id": "monthly-close-detail",
                "format": "csv",
                "parameters": {"period": "2026-07"},
            },
            headers=viewer_headers,
        ).get_json()["result"]["structuredContent"]
        export_job_id = export_created["job"]["id"]

        def _succeeded():
            status = _call_tool(
                hosted_client,
                "export_get",
                {"job_id": export_job_id},
                headers=viewer_headers,
            ).get_json()["result"]["structuredContent"]
            return status if status["job"]["status"] == "succeeded" else None

        export_status = wait_for(_succeeded, timeout=5, message="export job to succeed")
        download_link = _call_tool(
            hosted_client,
            "export_download_link_create",
            {"job_id": export_job_id},
            headers=viewer_headers,
        ).get_json()["result"]["structuredContent"]["download_url"]
        hosted_app.extensions["registry"].revoke_app_access(
            "hosted-outputs",
            principal_type="user",
            principal_id="trusted_proxy:viewer-1",
        )
        revoked_download = hosted_client.get(download_link, headers=viewer_headers)

        return {
            "created_is_error": created.get_json()["result"]["isError"],
            "mcp_status": mcp_response.status_code,
            "mcp_output_count": mcp_response.get_json()["result"]["structuredContent"]["output_count"],
            "ui_status": ui_response.status_code,
            "ui_data": ui_response.data,
            "denied_mcp_status": denied_mcp.status_code,
            "denied_mcp_category": denied_mcp.get_json()["error"]["data"]["category"],
            "denied_ui_status": denied_ui.status_code,
            "denied_ui_category": denied_ui.get_json()["error"]["data"]["category"],
            "export_job_status": export_status["job"]["status"],
            "download_link": download_link,
            "revoked_status": revoked_download.status_code,
            "revoked_category": revoked_download.get_json()["error"]["data"]["category"],
        }

    def test_viewer_sees_output_catalog_via_mcp_and_ui(self, flow):
        assert flow["created_is_error"] is False
        assert flow["mcp_status"] == 200
        assert flow["mcp_output_count"] == 1
        assert flow["ui_status"] == 200
        assert b"Monthly close detail" in flow["ui_data"]

    def test_stranger_is_denied_catalog_via_mcp_and_ui(self, flow):
        assert flow["denied_mcp_status"] == 403
        assert flow["denied_mcp_category"] == "mcp_authorization_denied"
        assert flow["denied_ui_status"] == 403
        assert flow["denied_ui_category"] == "consumption_authorization_denied"

    def test_viewer_can_export_and_receive_download_link(self, flow):
        assert flow["export_job_status"] == "succeeded"
        assert flow["download_link"]

    def test_revoked_viewer_download_is_denied(self, flow):
        assert flow["revoked_status"] == 403
        assert flow["revoked_category"] == "consumption_authorization_denied"


@pytest.mark.slow
def test_phase1_mcp_csv_export_is_pinned_encrypted_and_downloadable(app, client):
    _create_output_app(app, client)
    executor = _FakeDatasetExecutor()
    service = _enable_phase1(app, executor)

    created = _call_tool(
        client,
        "app_export_create",
        {
            "name": "finance-outputs",
            "output_id": "monthly-close-detail",
            "format": "csv",
            "parameters": {"period": "2026-07"},
            "idempotency_key": "mcp-export-1",
        },
    ).get_json()["result"]["structuredContent"]
    job_id = created["job"]["id"]
    completed = _wait_for_job(client, job_id)
    resource_payload = _read_resource(client, f"dash://exports/{job_id}")

    assert completed["job"]["status"] == "succeeded"
    assert completed["job"]["revision_number"] == 1
    assert completed["job"]["output_contract_hash"]
    assert completed["job"]["policy_version"] == service.policy_version
    assert "parameters" not in completed["job"]
    assert completed["artifact"]["row_count"] == 2
    assert resource_payload == {
        key: value for key, value in completed.items() if key != "guidance"
    }
    assert executor.preflight_calls == 1

    with sqlite3.connect(app.config["REGISTRY_DB_PATH"]) as connection:
        encoded = connection.execute("SELECT parameters_json FROM consumption_jobs WHERE id = ?", (job_id,)).fetchone()[
            0
        ]
    assert "2026-07" not in encoded
    assert service.parameter_codec.decode(encoded) == {"period": "2026-07"}

    link = _call_tool(client, "export_download_link_create", {"job_id": job_id}).get_json()["result"][
        "structuredContent"
    ]
    downloaded = client.get(link["download_url"])
    assert downloaded.status_code == 200
    assert downloaded.headers["Cache-Control"] == "private, no-store"
    assert downloaded.headers["X-Content-Type-Options"] == "nosniff"
    assert downloaded.data.startswith(b"PERIOD,VALUE\n")
    assert b"'=SUM(A1:A2)" in downloaded.data

    artifact_path = service.artifact_store.resolve(completed["artifact"]["storage_key"])
    with sqlite3.connect(app.config["REGISTRY_DB_PATH"]) as connection:
        connection.execute(
            "UPDATE consumption_artifacts SET expires_at = ? WHERE job_id = ?",
            ("2000-01-01T00:00:00Z", job_id),
        )
        connection.commit()
    assert service.cleanup_expired_artifacts() == 1
    assert artifact_path.exists() is False
    expired = _call_tool(client, "export_get", {"job_id": job_id})
    assert expired.get_json()["result"]["structuredContent"]["job"]["status"] == "expired"
    with sqlite3.connect(app.config["REGISTRY_DB_PATH"]) as connection:
        audit_events = {
            row[0]
            for row in connection.execute(
                "SELECT event_type FROM consumption_audit_events WHERE job_id = ?", (job_id,)
            ).fetchall()
        }
    assert {
        "export.created",
        "export.succeeded",
        "artifact.downloaded",
        "artifact.expired",
    } <= audit_events


@pytest.mark.slow
def test_phase1_idempotency_and_parameter_validation(app, client):
    _create_output_app(app, client)
    _enable_phase1(app)
    request = {
        "name": "finance-outputs",
        "output_id": "monthly-close-detail",
        "format": "csv",
        "parameters": {"period": "2026-07"},
        "idempotency_key": "same-request",
    }
    first = _call_tool(client, "app_export_create", request).get_json()["result"]
    second = _call_tool(client, "app_export_create", request).get_json()["result"]
    assert first["structuredContent"]["job"]["id"] == second["structuredContent"]["job"]["id"]

    conflict = _call_tool(
        client,
        "app_export_create",
        {**request, "parameters": {"period": "2026-08"}},
    ).get_json()["result"]
    assert conflict["isError"] is True
    assert conflict["structuredContent"]["error"]["category"] == "consumption_idempotency_conflict"

    invalid = _call_tool(
        client,
        "app_export_create",
        {**request, "idempotency_key": "invalid-params", "parameters": {"period": "July"}},
    ).get_json()["result"]
    assert invalid["isError"] is True
    assert invalid["structuredContent"]["error"]["category"] == "consumption_parameter_validation_error"


@pytest.mark.slow
def test_phase1_ui_uses_csrf_and_same_job_service(app, client):
    _create_output_app(app, client)
    _enable_phase1(app)
    page = client.get("/manage/apps/finance-outputs/consumption")
    assert page.status_code == 200
    csrf_match = re.search(rb'name="_csrf" value="([^"]+)"', page.data)
    idempotency_match = re.search(rb'name="idempotency_key" value="([^"]+)"', page.data)
    assert csrf_match and idempotency_match

    denied = client.post(
        "/manage/apps/finance-outputs/exports",
        data={
            "output_id": "monthly-close-detail",
            "format": "csv",
            "param__period": "2026-07",
            "type__period": "string",
            "idempotency_key": "ui-denied",
        },
    )
    assert denied.status_code == 403
    assert denied.get_json()["error"]["data"]["category"] == "consumption_token_invalid"

    created = client.post(
        "/manage/apps/finance-outputs/exports",
        data={
            "_csrf": csrf_match.group(1).decode(),
            "output_id": "monthly-close-detail",
            "format": "csv",
            "param__period": "2026-07",
            "type__period": "string",
            "idempotency_key": idempotency_match.group(1).decode(),
        },
    )
    assert created.status_code == 303
    job_id = created.headers["Location"].rsplit("/", 1)[-1]
    completed = _wait_for_job(client, job_id)
    assert completed["job"]["status"] == "succeeded"
    mcp_same = _call_tool(
        client,
        "app_export_create",
        {
            "name": "finance-outputs",
            "output_id": "monthly-close-detail",
            "format": "csv",
            "parameters": {"period": "2026-07"},
            "idempotency_key": idempotency_match.group(1).decode(),
        },
    ).get_json()["result"]["structuredContent"]
    assert mcp_same["job"]["id"] == job_id
    detail = client.get(created.headers["Location"])
    assert detail.status_code == 200
    assert b"Download CSV" in detail.data


@pytest.mark.slow
def test_phase1_cancellation_publishes_no_partial_artifact(app, client):
    _create_output_app(app, client)
    executor = _BlockingDatasetExecutor()
    service = _enable_phase1(app, executor)
    created = _call_tool(
        client,
        "app_export_create",
        {
            "name": "finance-outputs",
            "output_id": "monthly-close-detail",
            "format": "csv",
            "parameters": {"period": "2026-07"},
        },
    ).get_json()["result"]["structuredContent"]
    job_id = created["job"]["id"]
    assert executor.started.wait(timeout=2)
    cancelled = _call_tool(client, "export_cancel", {"job_id": job_id})
    assert cancelled.status_code == 200
    executor.release.set()
    terminal = _wait_for_job(client, job_id)
    assert terminal["job"]["status"] == "cancelled"
    assert terminal["artifact"] is None
    assert not list((service.artifact_store.root / job_id).glob("*.csv"))


@pytest.mark.slow
def test_phase1_row_limit_fails_without_artifact(app, client):
    _create_output_app(app, client)
    service = _enable_phase1(app, _FakeDatasetExecutor())
    service.policy = replace(service.policy, max_rows=1)
    service.policy_version = service.policy.version
    created = _call_tool(
        client,
        "app_export_create",
        {
            "name": "finance-outputs",
            "output_id": "monthly-close-detail",
            "format": "csv",
            "parameters": {"period": "2026-07"},
        },
    ).get_json()["result"]["structuredContent"]
    terminal = _wait_for_job(client, created["job"]["id"])
    assert terminal["job"]["status"] == "failed"
    assert terminal["job"]["error"]["category"] == "consumption_export_limit_exceeded"
    assert terminal["artifact"] is None


def test_exasol_export_executor_uses_bounded_fetch_and_closes_resources(tmp_path: Path):
    artifact_root = tmp_path / "artifact"
    (artifact_root / "queries").mkdir(parents=True)
    (artifact_root / "queries" / "export.sql").write_text("SELECT {period!s} AS PERIOD FROM DUAL", encoding="utf-8")

    class Statement:
        def __init__(self) -> None:
            self.batches = [[("2026-07",), ("2026-08",)], []]
            self.fetch_sizes: list[int] = []
            self.closed = False

        def column_names(self):
            return ["PERIOD"]

        def fetchmany(self, size):
            self.fetch_sizes.append(size)
            return self.batches.pop(0)

        def fetchall(self):
            raise AssertionError("export execution must not call fetchall")

        def close(self):
            self.closed = True

    class Connection:
        def __init__(self, statement) -> None:
            self.statement = statement
            self.closed = False

        def execute(self, sql, parameters):
            assert parameters == {"period": "2026-07"}
            return self.statement

        def close(self):
            self.closed = True

    statement = Statement()
    connection = Connection(statement)
    profile = SimpleNamespace(name="analytics-prod")
    connection_options = {}

    def connect_uncached(_profile, **kwargs):
        connection_options.update(kwargs)
        return connection

    service = SimpleNamespace(
        profile_store=SimpleNamespace(get_profile=lambda _name: profile),
        connection_manager=SimpleNamespace(connect_uncached=connect_uncached),
    )
    executor = ExasolDatasetExecutor(service, batch_size=2, max_runtime_seconds=30)
    revision = SimpleNamespace(
        app_name="finance-outputs",
        revision_number=1,
        artifact_path=str(artifact_root),
        manifest={"data_sources": {"primary": {"kind": "exasol", "profile": "analytics-prod"}}},
    )
    output = {
        "id": "monthly-close-detail",
        "source": {
            "type": "exasol_sql",
            "data_source": "primary",
            "path": "queries/export.sql",
        },
    }

    stream = executor.stream(
        revision,
        output,
        {"period": "2026-07"},
        cancelled=lambda: False,
    )
    assert stream.columns == ["PERIOD"]
    assert list(stream.batches) == [[["2026-07"], ["2026-08"]]]
    assert statement.fetch_sizes == [2, 2]
    assert connection_options == {"query_timeout_seconds": 30}
    assert statement.closed is True
    assert connection.closed is True


# ---------------------------------------------------------------------------
# Phase 2 — XLSX, durable job center, quotas, reconciliation
# ---------------------------------------------------------------------------


class _FlakyDatasetExecutor(_FakeDatasetExecutor):
    """Fails the first stream call with an unexpected error, then succeeds."""

    def __init__(self) -> None:
        super().__init__()
        self.stream_calls = 0

    def stream(self, revision, output, parameters, *, cancelled) -> DatasetStream:
        self.stream_calls += 1
        if self.stream_calls == 1:
            raise RuntimeError("transient connection failure")
        return super().stream(revision, output, parameters, cancelled=cancelled)


def _wait_store_status(store, job_id: str, expected: set[str], timeout: float = 5.0):
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        job = store.get_job(job_id)
        if job is not None and job.status in expected:
            return job
        time.sleep(0.02)
    raise AssertionError(f"Job {job_id} did not reach {expected}")


def test_phase2_csv_and_xlsx_writers_pass_golden_files(tmp_path: Path):
    from dash_server.consumption.csv_format import write_csv
    from dash_server.consumption.xlsx_format import write_xlsx
    from openpyxl import load_workbook

    columns = ["PERIOD", "COUNT", "RATIO", "NOTE", "FORMULA", "NEGATIVE", "EMPTY"]
    rows = [
        ["2026-07", 42, 0.125, "plain text", "=SUM(A1:A2)", "-lead", None],
        ["2026-08", 7, 2.5, "comma, text", "@cmd", "+1+1", None],
    ]

    csv_path = tmp_path / "export.csv"
    csv_result = write_csv(
        csv_path,
        columns=columns,
        batches=iter([rows]),
        max_rows=100,
        max_bytes=100_000,
        cancelled=lambda: False,
    )
    golden = Path(__file__).parent / "golden" / "consumption_export.csv"
    assert csv_path.read_bytes() == golden.read_bytes()
    assert csv_result["row_count"] == 2

    xlsx_path = tmp_path / "export.xlsx"
    xlsx_result = write_xlsx(
        xlsx_path,
        columns=columns,
        batches=iter([rows]),
        max_rows=100,
        max_bytes=1_000_000,
        cancelled=lambda: False,
        provenance={"app": "finance-outputs", "output_id": "monthly-close-detail"},
    )
    assert xlsx_result["row_count"] == 2
    workbook = load_workbook(xlsx_path)
    assert workbook.sheetnames == ["monthly-close-detail", "Provenance"]
    sheet = workbook["monthly-close-detail"]
    assert sheet.freeze_panes == "A2"
    typed = [[(cell.value, cell.data_type) for cell in row] for row in sheet.iter_rows(min_row=2)]
    assert typed == [
        [
            ("2026-07", "s"),
            (42, "n"),
            (0.125, "n"),
            ("plain text", "s"),
            ("=SUM(A1:A2)", "s"),  # a formula from source data must stay a string
            ("-lead", "s"),
            (None, "n"),
        ],
        [
            ("2026-08", "s"),
            (7, "n"),
            (2.5, "n"),
            ("comma, text", "s"),
            ("@cmd", "s"),
            ("+1+1", "s"),
            (None, "n"),
        ],
    ]


@pytest.mark.slow
def test_phase2_xlsx_export_end_to_end_with_provenance(app, client):
    from io import BytesIO

    from openpyxl import load_workbook

    _create_output_app(app, client)
    _enable_phase1(app)

    outputs = _call_tool(client, "app_outputs_list", {"name": "finance-outputs"}).get_json()["result"][
        "structuredContent"
    ]
    policy = outputs["outputs"][0]["policy"]
    assert policy["phase"] == "on_demand_exports"
    assert policy["format_availability"] == {
        "csv": {"executable": True, "reason": "available"},
        "xlsx": {"executable": True, "reason": "available"},
    }
    assert policy["executable"] is True
    assert policy["reason"] == "available"

    page = client.get("/manage/apps/finance-outputs/consumption")
    assert page.status_code == 200
    assert b'<option value="csv">CSV</option>' in page.data
    assert b'<option value="xlsx">XLSX</option>' in page.data

    created = _call_tool(
        client,
        "app_export_create",
        {
            "name": "finance-outputs",
            "output_id": "monthly-close-detail",
            "format": "xlsx",
            "parameters": {"period": "2026-07"},
        },
    ).get_json()["result"]["structuredContent"]
    payload = _wait_for_job(client, created["job"]["id"])
    assert payload["job"]["status"] == "succeeded"
    assert payload["artifact"]["filename"].endswith(".xlsx")
    assert payload["artifact"]["content_type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    link = _call_tool(client, "export_download_link_create", {"job_id": created["job"]["id"]}).get_json()["result"][
        "structuredContent"
    ]
    downloaded = client.get(link["download_url"])
    assert downloaded.status_code == 200
    workbook = load_workbook(BytesIO(downloaded.data))
    sheet = workbook["monthly-close-detail"]
    cells = {(cell.row, cell.column): cell for row in sheet.iter_rows() for cell in row}
    assert cells[(2, 2)].value == "=SUM(A1:A2)"
    assert cells[(2, 2)].data_type == "s"
    provenance = {row[0].value: row[1].value for row in workbook["Provenance"].iter_rows()}
    assert provenance["app"] == "finance-outputs"
    assert provenance["output_id"] == "monthly-close-detail"
    assert provenance["parameters"] == '{"period": "<provided>"}'
    assert provenance["limit_outcome"] == "within_limits"
    assert "2026-07" not in provenance["parameters"]


@pytest.mark.slow
def test_phase2_retry_recovers_from_transient_failure(app, client):
    _create_output_app(app, client)
    executor = _FlakyDatasetExecutor()
    _enable_phase1(app, executor)

    created = _call_tool(
        client,
        "app_export_create",
        {
            "name": "finance-outputs",
            "output_id": "monthly-close-detail",
            "format": "csv",
            "parameters": {"period": "2026-07"},
        },
    ).get_json()["result"]["structuredContent"]
    payload = _wait_for_job(client, created["job"]["id"])

    assert payload["job"]["status"] == "succeeded"
    assert payload["job"]["attempt_count"] == 2
    assert executor.stream_calls == 2
    db_path = app.config["REGISTRY_DB_PATH"]
    with sqlite3.connect(db_path) as connection:
        events = {
            row[0]
            for row in connection.execute(
                "SELECT event_type FROM consumption_audit_events WHERE job_id = ?",
                (created["job"]["id"],),
            ).fetchall()
        }
    assert {"export.created", "export.retried", "export.succeeded"} <= events


@pytest.mark.slow
def test_phase2_restart_reconciliation_strands_nothing(app, client):
    from dash_server.consumption.service import ConsumptionService

    _create_output_app(app, client)
    _enable_phase1(app)
    job_ids = []
    for index in range(3):
        created = _call_tool(
            client,
            "app_export_create",
            {
                "name": "finance-outputs",
                "output_id": "monthly-close-detail",
                "format": "csv",
                "parameters": {"period": "2026-07"},
                "idempotency_key": f"restart-{index}",
            },
        ).get_json()["result"]["structuredContent"]
        job_ids.append(created["job"]["id"])
        _wait_for_job(client, created["job"]["id"])

    queued_id, retryable_id, exhausted_id = job_ids
    db_path = app.config["REGISTRY_DB_PATH"]
    with sqlite3.connect(db_path) as connection:
        connection.execute(
            "UPDATE consumption_jobs SET status='queued', finished_at=NULL, error_json=NULL, "
            "lease_owner=NULL, lease_expires_at=NULL, attempt_count=0 WHERE id=?",
            (queued_id,),
        )
        connection.execute(
            "UPDATE consumption_jobs SET status='running', finished_at=NULL, error_json=NULL, "
            "lease_owner='dead-process', lease_expires_at='2020-01-01T00:00:00Z', attempt_count=1 WHERE id=?",
            (retryable_id,),
        )
        connection.execute(
            "UPDATE consumption_jobs SET status='running', finished_at=NULL, error_json=NULL, "
            "lease_owner='dead-process', lease_expires_at='2020-01-01T00:00:00Z', attempt_count=2 WHERE id=?",
            (exhausted_id,),
        )
        connection.commit()

    restarted = ConsumptionService(
        app.extensions["registry"],
        app.extensions["authorization_service"],
        dict(app.config),
        exasol_service=None,
        artifacts_root=app.config["ARTIFACTS_ROOT"],
    )
    restarted.executor = _FakeDatasetExecutor()
    restarted.policy = replace(restarted.policy, exports_enabled=True)
    restarted.start()

    requeued = _wait_store_status(restarted.store, queued_id, {"succeeded"})
    retried = _wait_store_status(restarted.store, retryable_id, {"succeeded"})
    stranded = _wait_store_status(restarted.store, exhausted_id, {"failed"})
    assert requeued.status == "succeeded"
    assert retried.status == "succeeded"
    assert stranded.error is not None and stranded.error["category"] == "consumption_job_stranded"
    assert restarted.store.list_incomplete_jobs() == []
    with sqlite3.connect(db_path) as connection:
        events = {
            row[0]
            for row in connection.execute(
                "SELECT event_type FROM consumption_audit_events WHERE job_id = ?",
                (retryable_id,),
            ).fetchall()
        }
    assert "export.requeued" in events


@pytest.mark.slow
def test_phase2_policy_tightening_narrows_pinned_job_limits(app, client):
    _create_output_app(app, client)
    service = _enable_phase1(app)
    held_jobs: list[str] = []
    service.coordinator.submit = held_jobs.append

    created = _call_tool(
        client,
        "app_export_create",
        {
            "name": "finance-outputs",
            "output_id": "monthly-close-detail",
            "format": "csv",
            "parameters": {"period": "2026-07"},
        },
    ).get_json()["result"]["structuredContent"]
    job_id = created["job"]["id"]
    assert held_jobs == [job_id]
    assert created["job"]["effective_limits"]["max_rows"] == 5000

    service.policy = replace(service.policy, max_rows=1)
    service._run_job(job_id)

    job = service.store.get_job(job_id)
    assert job is not None and job.status == "failed"
    assert job.error is not None and job.error["category"] == "consumption_export_limit_exceeded"
    assert job.effective_limits["max_rows"] == 5000  # pinned limits stayed pinned; the applied cap narrowed
    assert service.store.get_artifact_for_job(job_id) is None


@pytest.mark.slow
def test_phase2_quotas_bound_active_jobs(app, client):
    _create_output_app(app, client)
    service = _enable_phase1(app)
    service.coordinator.submit = lambda job_id: None
    service.policy = replace(service.policy, max_active_jobs_per_principal=2, max_active_jobs_per_app=10)

    for index in range(2):
        response = _call_tool(
            client,
            "app_export_create",
            {
                "name": "finance-outputs",
                "output_id": "monthly-close-detail",
                "format": "csv",
                "parameters": {"period": "2026-07"},
                "idempotency_key": f"quota-{index}",
            },
        )
        assert response.status_code == 200

    principal_blocked = _call_tool(
        client,
        "app_export_create",
        {
            "name": "finance-outputs",
            "output_id": "monthly-close-detail",
            "format": "csv",
            "parameters": {"period": "2026-07"},
            "idempotency_key": "quota-overflow",
        },
    )
    principal_result = principal_blocked.get_json()["result"]
    assert principal_result["isError"] is True
    assert principal_result["structuredContent"]["error"]["category"] == "consumption_quota_exceeded"
    assert principal_result["structuredContent"]["error"]["details"]["scope"] == "principal"

    service.policy = replace(service.policy, max_active_jobs_per_principal=10, max_active_jobs_per_app=2)
    app_blocked = _call_tool(
        client,
        "app_export_create",
        {
            "name": "finance-outputs",
            "output_id": "monthly-close-detail",
            "format": "csv",
            "parameters": {"period": "2026-07"},
            "idempotency_key": "quota-overflow",
        },
    )
    app_result = app_blocked.get_json()["result"]
    assert app_result["isError"] is True
    assert app_result["structuredContent"]["error"]["category"] == "consumption_quota_exceeded"
    assert app_result["structuredContent"]["error"]["details"]["scope"] == "app"

    # Finishing the held jobs releases quota.
    for job in service.store.list_incomplete_jobs():
        service._run_job(job.id)
    released = _call_tool(
        client,
        "app_export_create",
        {
            "name": "finance-outputs",
            "output_id": "monthly-close-detail",
            "format": "csv",
            "parameters": {"period": "2026-07"},
            "idempotency_key": "quota-after-release",
        },
    )
    assert released.status_code == 200
    assert released.get_json()["result"]["isError"] is False


@pytest.mark.slow
def test_phase2_retention_prunes_jobs_and_releases_idempotency_keys(app, client):
    _create_output_app(app, client)
    service = _enable_phase1(app)

    created = _call_tool(
        client,
        "app_export_create",
        {
            "name": "finance-outputs",
            "output_id": "monthly-close-detail",
            "format": "csv",
            "parameters": {"period": "2026-07"},
            "idempotency_key": "retained-key",
        },
    ).get_json()["result"]["structuredContent"]
    job_id = created["job"]["id"]
    _wait_for_job(client, job_id)
    artifact = service.store.get_artifact_for_job(job_id)
    assert artifact is not None
    artifact_path = service.artifact_store.resolve(artifact.storage_key)

    db_path = app.config["REGISTRY_DB_PATH"]
    with sqlite3.connect(db_path) as connection:
        connection.execute(
            "UPDATE consumption_jobs SET finished_at='2020-01-01T00:00:00Z' WHERE id=?",
            (job_id,),
        )
        connection.execute(
            "INSERT INTO consumption_audit_events "
            "(event_type, actor_principal_id, app_name, decision, details_json, created_at) "
            "VALUES ('export.created', 'p', 'finance-outputs', 'allowed', '{}', '2019-01-01T00:00:00Z')"
        )
        connection.commit()

    summary = service.run_maintenance()

    assert summary["pruned_jobs"] == 1
    assert service.store.get_job(job_id) is None
    assert not artifact_path.exists()
    with sqlite3.connect(db_path) as connection:
        stale_audit = connection.execute(
            "SELECT COUNT(*) FROM consumption_audit_events WHERE created_at <= '2019-12-31'"
        ).fetchone()[0]
    assert stale_audit == 0

    reused = _call_tool(
        client,
        "app_export_create",
        {
            "name": "finance-outputs",
            "output_id": "monthly-close-detail",
            "format": "csv",
            "parameters": {"period": "2026-07"},
            "idempotency_key": "retained-key",
        },
    )
    assert reused.status_code == 200
    assert reused.get_json()["result"]["structuredContent"]["job"]["id"] != job_id


@pytest.mark.slow
def test_phase2_schema_ledger_records_versions_and_refuses_downgrade(app):
    service = app.extensions["consumption_service"]
    db_path = app.config["REGISTRY_DB_PATH"]
    with sqlite3.connect(db_path) as connection:
        versions = {
            row[0] for row in connection.execute("SELECT version FROM consumption_schema_migrations").fetchall()
        }
    assert versions == {1, 2}

    with sqlite3.connect(db_path) as connection:
        connection.execute(
            "INSERT INTO consumption_schema_migrations (version, applied_at) VALUES (99, '2099-01-01T00:00:00Z')"
        )
        connection.commit()
    with pytest.raises(RuntimeError, match="newer than this server supports"):
        service.store.initialize()


@pytest.mark.slow
def test_phase2_multi_process_coordinator_refusal_and_status(app, client):
    service = app.extensions["consumption_service"]

    with pytest.raises(RuntimeError, match="single-process only"):
        service.store.claim_coordinator(
            owner="other-process",
            pid=999_999,
            stale_after_seconds=300,
            is_pid_alive=lambda pid: True,
        )

    status = service.coordinator_status()
    assert status["mode"] == "local-single-process"
    assert status["multi_process_supported"] is False
    assert status["claim"]["owner"] == service.instance_id

    runtime_status = _read_resource(client, "dash://runtime/status")
    assert runtime_status["consumption_coordinator"]["mode"] == "local-single-process"
    assert runtime_status["consumption_coordinator"]["multi_process_supported"] is False

    # A dead previous owner can be taken over.
    service.store.claim_coordinator(
        owner="successor",
        pid=999_999,
        stale_after_seconds=300,
        is_pid_alive=lambda pid: False,
    )
    assert service.store.coordinator_snapshot()["owner"] == "successor"
    # Restore the fixture's claim for any later maintenance calls.
    service.store.claim_coordinator(
        owner=service.instance_id,
        pid=999_999,
        stale_after_seconds=300,
        is_pid_alive=lambda pid: False,
    )


@pytest.mark.slow
def test_phase2_admin_job_view_requires_capability_and_redacts(tmp_path: Path):
    hosted_app = create_app(
        {
            "TESTING": True,
            "REGISTRY_DB_PATH": str(tmp_path / "registry.sqlite3"),
            "ARTIFACTS_ROOT": str(tmp_path / "artifacts"),
            "WORKSPACES_ROOT": str(tmp_path / "workspaces"),
            "DIAGNOSTICS_ROOT": str(tmp_path / "diagnostics"),
            "DEPENDENCY_STATE_ROOT": str(tmp_path / "dependency_state"),
            "GITOPS_REPO_PATH": str(tmp_path / "gitops-repo"),
            "EXASOL_SECRETS_ROOT": str(tmp_path / "exasol-secrets"),
            "AUTO_INSTALL_DEPENDENCIES": False,
            "PYTHON_EXECUTABLE": sys.executable,
            "DASH_SERVER_MODE": "hosted",
            "SECRET_KEY": "test-secret-key",
            "SESSION_COOKIE_SECURE": True,
            "SESSION_COOKIE_HTTPONLY": True,
            "SESSION_COOKIE_SAMESITE": "Lax",
            "DASH_SERVER_PUBLIC_BASE_URL": "https://dash.example.test",
            "DASH_SERVER_AUTH_PROVIDER": "trusted_proxy",
            "DASH_SERVER_TRUSTED_PROXY_HEADERS_ENABLED": True,
            "DASH_SERVER_TRUSTED_PROXY_ALLOWED_CIDRS": ("127.0.0.1/32",),
            "DASH_SERVER_BOOTSTRAP_ADMIN_PRINCIPAL_IDS": ("trusted_proxy:admin-1",),
            "DASH_SERVER_ALLOW_UNSAFE_INPROCESS": True,
        }
    )
    hosted_client = hosted_app.test_client()
    admin_headers = {"X-Forwarded-User": "admin-1", "X-Forwarded-Email": "admin@example.test"}
    viewer_headers = {"X-Forwarded-User": "viewer-1", "X-Forwarded-Email": "viewer@example.test"}
    _wire_fake_exasol_connector(hosted_app)
    profile_created = _call_tool(
        hosted_client,
        "exasol_profile_create_local",
        {
            "name": "analytics-prod",
            "backend": "onprem",
            "credential_mode": "password",
            "dsn": "demodb.exasol.com:8563",
            "user": "sys",
            "secret_value": "super-secret",
        },
        headers=admin_headers,
    )
    assert profile_created.status_code == 200
    created = _call_tool(
        hosted_client,
        "app_create_from_files",
        {
            "name": "hosted-outputs",
            "data_sources": {"primary": {"kind": "exasol", "profile": "analytics-prod"}},
            "consumption": _consumption_contract(),
            "files": [
                {"path": "app.py", "content": _APP_PY},
                {"path": "queries/export.sql", "content": "SELECT {period!s} AS PERIOD FROM DUAL\n"},
                {
                    "path": "queries/sql_smoke.json",
                    "content": json.dumps({"queries/export.sql": {"period": "2026-07"}}) + "\n",
                },
            ],
        },
        headers=admin_headers,
    )
    assert created.get_json()["result"]["isError"] is False
    hosted_app.extensions["registry"].grant_app_access(
        "hosted-outputs",
        principal_type="user",
        principal_id="trusted_proxy:viewer-1",
        role="viewer",
        scope="all",
        created_by_principal_id="trusted_proxy:admin-1",
    )
    _enable_phase1(hosted_app)

    export_created = _call_tool(
        hosted_client,
        "app_export_create",
        {
            "name": "hosted-outputs",
            "output_id": "monthly-close-detail",
            "format": "csv",
            "parameters": {"period": "1999-12"},
        },
        headers=viewer_headers,
    ).get_json()["result"]["structuredContent"]
    job_id = export_created["job"]["id"]
    deadline = time.monotonic() + 5
    while time.monotonic() < deadline:
        status = _call_tool(hosted_client, "export_get", {"job_id": job_id}, headers=viewer_headers).get_json()[
            "result"
        ]["structuredContent"]
        if status["job"]["status"] == "succeeded":
            break
        time.sleep(0.02)
    assert status["job"]["status"] == "succeeded"

    viewer_denied = _call_tool(
        hosted_client,
        "app_exports_admin_list",
        {"name": "hosted-outputs"},
        headers=viewer_headers,
    )
    assert viewer_denied.status_code == 403
    assert viewer_denied.get_json()["error"]["data"]["category"] == "mcp_authorization_denied"
    viewer_denied_ui = hosted_client.get(
        "/manage/apps/hosted-outputs/consumption/jobs",
        headers=viewer_headers,
    )
    assert viewer_denied_ui.status_code == 403

    admin_payload = _call_tool(
        hosted_client,
        "app_exports_admin_list",
        {"name": "hosted-outputs"},
        headers=admin_headers,
    ).get_json()["result"]["structuredContent"]
    assert admin_payload["job_count"] == 1
    assert admin_payload["jobs"][0]["job"]["id"] == job_id
    assert admin_payload["jobs"][0]["job"]["requested_by_principal_id"] == "trusted_proxy:viewer-1"
    assert admin_payload["jobs"][0]["parameters_redacted"] == {"period": "<provided>"}
    assert "1999-12" not in json.dumps(admin_payload)
    assert admin_payload["coordinator"]["mode"] == "local-single-process"

    admin_ui = hosted_client.get(
        "/manage/apps/hosted-outputs/consumption/jobs",
        headers=admin_headers,
    )
    assert admin_ui.status_code == 200
    assert b"monthly-close-detail" in admin_ui.data
    assert b"1999-12" not in admin_ui.data
    assert "provided" in admin_ui.get_data(as_text=True)


def test_phase2_artifact_store_interface_round_trip(tmp_path: Path):
    from dash_server.consumption.artifacts import (
        InMemoryObjectClient,
        LocalArtifactStore,
        ObjectStoreArtifactStore,
    )

    job_id = "0a1b2c3d-4e5f-6789-abcd-ef0123456789"
    stores = [
        LocalArtifactStore(tmp_path / "local"),
        ObjectStoreArtifactStore(InMemoryObjectClient(), tmp_path / "object-cache"),
    ]
    for store in stores:
        staged = store.temporary_path(job_id)
        staged.write_bytes(b"PERIOD\n2026-07\n")
        key = store.publish(job_id, staged, "export.csv")
        resolved = store.resolve(key)
        assert resolved.read_bytes() == b"PERIOD\n2026-07\n"
        store.delete(key)
        with pytest.raises(DashServerError):
            store.resolve(key)

        discarded = store.temporary_path(job_id)
        discarded.write_bytes(b"partial")
        store.discard(discarded)
        assert not discarded.exists()
